from __future__ import annotations

import re
from pathlib import Path

import openpyxl
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import transaction

from core.utils.jalali import PERSIAN_DIGITS_TRANS
from store.models import CartItem, Category, Order, Product, ProductFeature


DIGITS_TO_ASCII = str.maketrans(
    "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
    "01234567890123456789",
)


def _to_ascii_digits(text: str) -> str:
    return (text or "").translate(DIGITS_TO_ASCII)


def _to_persian_digits(text: str) -> str:
    return (text or "").translate(PERSIAN_DIGITS_TRANS)


def _infer_category_name(name: str) -> str:
    cleaned = (name or "").strip()
    if "فرپیتزا" in cleaned or "فر پیتزا" in cleaned:
        return "فر پیتزا"
    if "دیسپلی" in cleaned:
        return "دیسپلی"
    if "سرخ" in cleaned:
        return "سرخ‌کن"
    if "گریل" in cleaned:
        return "گریل"
    return "سایر"


def _build_description(*, name: str, category_name: str) -> str:
    return (
        f"«{name}» از دسته {category_name} بوده و برای استفاده در آشپزخانه‌های صنعتی، فست‌فودها و مجموعه‌های خدمات غذایی مناسب است.\n"
        "طراحی صنعتی، بدنه مقاوم و عملکرد پایدار در استفاده طولانی‌مدت.\n"
        "ساخت ایران."
    )


def _extract_features(name: str, category_name: str) -> list[tuple[str, str]]:
    text = (name or "").strip()
    ascii_text = _to_ascii_digits(text)

    features: list[tuple[str, str]] = [
        ("ساخت", "ایران"),
        ("دسته‌بندی", category_name),
        ("کاربری", "آشپزخانه صنعتی"),
    ]

    if category_name == "فر پیتزا":
        mouth = None
        m = re.search(r"دهانه\s*([0-9]{2,3})", ascii_text)
        if m:
            mouth = _to_persian_digits(m.group(1))
            features.append(("دهانه", f"{mouth} سانتی‌متر"))

        motor_pos = None
        if "موتوربغل" in text or "موتور بغل" in text:
            motor_pos = "بغل"
        elif "موتورپایین" in text or "موتور پایین" in text:
            motor_pos = "پایین"
        if motor_pos:
            features.append(("موقعیت موتور", motor_pos))

        if "ریلی" in text:
            features.append(("نوع", "ریلی"))

    elif category_name == "دیسپلی":
        m = re.search(r"([0-9]{2,3})", ascii_text)
        if m:
            length = _to_persian_digits(m.group(1))
            features.append(("طول", f"{length} سانتی‌متر"))
        features.append(("نوع", "دیسپلی"))

    elif category_name == "سرخ‌کن":
        control = None
        if "دیجیتال" in text:
            control = "دیجیتال"
        elif "انالوگ" in text or "آنالوگ" in text:
            control = "آنالوگ"
        if control:
            features.append(("کنترل", control))

        m = re.search(r"([0-9]+)\s*لگن", ascii_text)
        if m:
            pans = _to_persian_digits(m.group(1))
            features.append(("تعداد لگن", pans))

        features.append(("نوع", "سرخ‌کن"))

    elif category_name == "گریل":
        m = re.search(r"([0-9]{2,3})", ascii_text)
        if m:
            width = _to_persian_digits(m.group(1))
            features.append(("عرض", f"{width} سانتی‌متر"))

        grill_type = None
        if "روغنی" in text:
            grill_type = "روغنی"
        elif "ذغالی" in text:
            grill_type = "ذغالی"
        elif "ترکیبی" in text:
            grill_type = "ترکیبی"
        if grill_type:
            features.append(("نوع پخت", grill_type))

        m = re.search(r"([0-9]+)\s*میل", ascii_text)
        if m:
            thickness = _to_persian_digits(m.group(1))
            features.append(("ضخامت صفحه", f"{thickness} میلی‌متر"))

        features.append(("نوع", "گریل"))

    else:
        first_word = text.split()[0] if text.split() else ""
        if first_word:
            features.append(("نوع", first_word))

    # Remove duplicates while preserving order.
    seen: set[str] = set()
    unique: list[tuple[str, str]] = []
    for k, v in features:
        if k in seen:
            continue
        seen.add(k)
        unique.append((k, v))
    return unique


def _parse_price_toman(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        rial = int(value)
    else:
        text = str(value).strip()
        text = _to_ascii_digits(text)
        text = re.sub(r"[^\d]", "", text)
        rial = int(text or 0)
    return max(0, rial // 10)


class Command(BaseCommand):
    help = "حذف محصولات فعلی و وارد کردن محصولات از فایل pricing.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "--path",
            default=str(Path(getattr(settings, "BASE_DIR", Path.cwd())) / "pricing.xlsx"),
            help="مسیر فایل Excel (xlsx)",
        )
        parser.add_argument(
            "--wipe",
            action="store_true",
            help="تمام محصولات/دسته‌ها و سفارش‌ها را پاک می‌کند و سپس ایمپورت انجام می‌دهد.",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="فقط شمارش می‌کند و دیتابیس را تغییر نمی‌دهد.",
        )
        parser.add_argument("--limit", type=int, default=0, help="محدود کردن تعداد محصولات واردشده (۰ یعنی همه)")

    def handle(self, *args, **options):
        path = Path(options["path"])
        wipe = bool(options["wipe"])
        dry_run = bool(options["dry_run"])
        limit = int(options["limit"] or 0)

        if not path.exists():
            raise FileNotFoundError(f"فایل یافت نشد: {path}")

        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]

        started = False
        imported_rows: list[tuple[str, int]] = []

        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            c1 = (str(row[0]).strip() if len(row) > 0 and row[0] is not None else "")
            c2 = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
            c3 = row[2] if len(row) > 2 else None

            if not started:
                if c1 == "ردیف" and "لیست" in c2 and "قیمت" in (str(c3) if c3 is not None else ""):
                    started = True
                continue

            try:
                int(c1)
            except Exception:
                continue
            if not c2:
                continue

            price_toman = _parse_price_toman(c3)
            imported_rows.append((c2, price_toman))
            if limit and len(imported_rows) >= limit:
                break

        _safe_write(self, self.style.SUCCESS(f"تعداد ردیف‌های قابل ایمپورت: {len(imported_rows)}"))
        if dry_run:
            return

        with transaction.atomic():
            if wipe:
                CartItem.objects.all().delete()
                # Delete orders first to avoid PROTECT on OrderItem.product.
                Order.objects.all().delete()
                ProductFeature.objects.all().delete()
                Product.objects.all().delete()
                Category.objects.all().delete()

            categories_by_name: dict[str, Category] = {c.name: c for c in Category.objects.all()}
            created_products: list[Product] = []

            for name, price in imported_rows:
                category_name = _infer_category_name(name)
                category = categories_by_name.get(category_name)
                if not category:
                    category = Category.objects.create(name=category_name)
                    categories_by_name[category_name] = category

                product = Product.objects.create(
                    name=name,
                    description=_build_description(name=name, category_name=category_name),
                    price=int(price),
                    domain=category_name,
                    category=category,
                    brand="",
                    sku="",
                    tags="",
                )
                created_products.append(product)

                features = _extract_features(name, category_name)
                ProductFeature.objects.bulk_create(
                    [ProductFeature(product=product, name=k, value=v) for (k, v) in features],
                    ignore_conflicts=False,
                )

        _safe_write(self, self.style.SUCCESS(f"ایمپورت انجام شد: {len(imported_rows)} محصول"))


def _safe_write(command: BaseCommand, message: str) -> None:
    encoding = getattr(command.stdout, "encoding", None) or "utf-8"
    safe_message = message.encode(encoding, errors="backslashreplace").decode(encoding, errors="ignore")
    command.stdout.write(safe_message)
